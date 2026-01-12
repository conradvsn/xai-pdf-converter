#!/usr/bin/env python3
"""
Report Generator Module
"""

import re
import random
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

from src.config import OPENPYXL_AVAILABLE, logger
from src.analysis.pdf_analyzer import detect_images_in_pdf
from src.analysis.sensitive_info_detector import detect_sensitive_information

if OPENPYXL_AVAILABLE:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter


# ===== ANONYMIZATION FUNCTIONS =====

def _anonymize_value(original_value, info_type, anonymization_map):
    """
    Génère un alias anonymisé pour une valeur donnée.

    Args:
        original_value: Valeur originale à anonymiser
        info_type: Type d'information (email, phone, company_name, etc.)
        anonymization_map: Dictionnaire pour maintenir la cohérence des mappings

    Returns:
        Valeur anonymisée
    """
    # Si déjà anonymisé, retourner la valeur existante
    if original_value in anonymization_map:
        return anonymization_map[original_value]

    # Extraire les suffixes à préserver (Inc, LLC, Ltd, etc.)
    suffix_patterns = r'\s+(Inc\.?|LLC\.?|Ltd\.?|Corporation|Corp\.?|Company|Co\.?|LP|LLP|SA|SAS|SARL|GmbH|AG)(\s*,?\s*)$'
    suffix_match = re.search(suffix_patterns, original_value, re.IGNORECASE)
    suffix = suffix_match.group(0) if suffix_match else ''
    base_value = original_value[:len(original_value) - len(suffix)] if suffix else original_value

    anonymized = ''

    if info_type == 'company_name':
        # IMPORTANT: Vérifier si cette valeur est liée à un nom d'entreprise déjà anonymisé
        # Ex: Si "Fidelity" → "Apex", alors "Fidelity Inc" → "Apex Inc"
        found_base = None
        found_base_anon = None

        # Chercher la correspondance la plus longue dans le mapping existant (bidirectionnel)
        for existing_original, existing_anon in anonymization_map.items():
            # Extraire la base de l'original existant (sans suffixes)
            existing_base = re.sub(r'\s+(Inc\.?|LLC\.?|Ltd\.?|Corporation|Corp\.?|Company|Co\.?|LP|LLP|SA|SAS|SARL|GmbH|AG)(\s*,?\s*)$', '', existing_original, flags=re.IGNORECASE).strip()

            # CAS 1: L'original existant est contenu dans la nouvelle valeur
            # Ex: existing="Fifth Third" in new="Fifth Third Bank"
            if existing_base in base_value and len(existing_base) > len(found_base or ''):
                found_base = existing_base
                # Extraire l'alias de base (sans suffixes Inc, LLC, etc.)
                found_base_anon = re.sub(r'\s+(Inc\.?|LLC\.?|Ltd\.?|Corporation|Corp\.?|Company|Co\.?|LP|LLP|SA|SAS|SARL|GmbH|AG)(\s*,?\s*)$', '', existing_anon, flags=re.IGNORECASE).strip()

            # CAS 2: La nouvelle valeur est contenue dans l'original existant
            # Ex: new="Fifth Third" in existing="Fifth Third Bancorp"
            elif base_value in existing_base and len(base_value) > len(found_base or ''):
                found_base = base_value
                # Utiliser le même alias de base
                found_base_anon = re.sub(r'\s+(Inc\.?|LLC\.?|Ltd\.?|Corporation|Corp\.?|Company|Co\.?|LP|LLP|SA|SAS|SARL|GmbH|AG)(\s*,?\s*)$', '', existing_anon, flags=re.IGNORECASE).strip()

        if found_base and found_base_anon:
            # Utiliser le même alias de base
            # Remplacer la partie commune et garder le reste
            remaining_part = base_value.replace(found_base, '').strip()

            if remaining_part:
                # Il y a des mots supplémentaires après le nom de base
                # Ex: "Fidelity Global Payments" → "Apex Global Payments"
                anonymized = f"{found_base_anon} {remaining_part}{suffix}"
            else:
                # Juste le nom de base avec suffix
                anonymized = f"{found_base_anon}{suffix}"
        else:
            # Nouveau nom d'entreprise, créer un nouvel alias
            # Générateur de noms d'entreprise réalistes et variés

            random.seed(hash(base_value))  # Seed pour cohérence

            # Stratégie 1: Noms réalistes de villes/régions + type d'entreprise (40%)
            city_names = [
                'Riverside', 'Oakmont', 'Clearwater', 'Fairview', 'Bridgeport', 'Summit',
                'Westfield', 'Highland', 'Lakeside', 'Greenfield', 'Parkview', 'Eastwood',
                'Mountainview', 'Oceanside', 'Hillcrest', 'Meadowbrook', 'Pinehurst', 'Brookfield',
                'Maplewood', 'Redwood', 'Silverstone', 'Northpoint', 'Southgate', 'Bayshore'
            ]
            business_types = [
                'Capital', 'Financial', 'Investments', 'Advisors', 'Partners', 'Associates',
                'Holdings', 'Trust', 'Securities', 'Asset Management', 'Wealth Management',
                'Equity', 'Ventures', 'Credit', 'Bancorp', 'Financial Group'
            ]

            # Stratégie 2: Noms de créateurs fictifs (patronymes) + type (30%)
            founder_names = [
                'Anderson', 'Bennett', 'Carter', 'Davidson', 'Edwards', 'Franklin', 'Goldman',
                'Harrison', 'Jensen', 'Kennedy', 'Lawrence', 'Morrison', 'Nelson', 'Patterson',
                'Reynolds', 'Stratton', 'Thompson', 'Vincent', 'Wallace', 'Wellington'
            ]
            firm_types = [
                'Group', 'Partners', '& Associates', 'Capital', 'Advisors', 'Holdings',
                'Financial', 'Investments', 'Securities', 'Asset Management', 'Trust'
            ]

            # Stratégie 3: Mots prestigieux + type d'entreprise (20%)
            prestige_words = [
                'Sterling', 'Premier', 'Crown', 'Royal', 'Pinnacle', 'Meridian', 'Horizon',
                'Crest', 'Legacy', 'Heritage', 'Signature', 'Cornerstone', 'Foundation',
                'Keystone', 'Milestone', 'Beacon', 'Gateway', 'Triumph', 'Victory', 'Liberty'
            ]

            # Stratégie 4: Combinaisons de 2 noms de fondateurs (10%)
            # Ex: "Morgan Stanley", "Goldman Sachs", "Merrill Lynch"

            # Choisir une stratégie
            strategy = random.randint(1, 10)

            if strategy <= 4:  # 40% - Ville + Type
                city = random.choice(city_names)
                biz_type = random.choice(business_types)
                anonymized = f"{city} {biz_type}{suffix}"

            elif strategy <= 7:  # 30% - Fondateur + Type
                founder = random.choice(founder_names)
                firm_type = random.choice(firm_types)
                if '&' in firm_type:
                    anonymized = f"{founder} {firm_type}{suffix}"
                else:
                    anonymized = f"{founder} {firm_type}{suffix}"

            elif strategy <= 9:  # 20% - Prestige + Type
                prestige = random.choice(prestige_words)
                biz_type = random.choice(business_types)
                anonymized = f"{prestige} {biz_type}{suffix}"

            else:  # 10% - Deux fondateurs
                founder1 = random.choice(founder_names)
                founder2 = random.choice([f for f in founder_names if f != founder1])
                firm_type = random.choice(['Group', 'Partners', 'Capital', 'Associates', 'Holdings'])
                anonymized = f"{founder1} {founder2} {firm_type}{suffix}"

    elif info_type == 'person_name':
        # Générateur de noms de personnes avec beaucoup plus de variété

        # IMPORTANT: Normaliser le nom pour la cohérence
        # "John Smith" et "J. Smith" doivent mapper au même alias
        normalized_for_hash = base_value.lower().strip()

        # Si c'est une initiale + nom, on cherche dans le mapping existant
        # pour voir si on a déjà le nom complet
        parts_check = base_value.split()
        potential_match = None

        if len(parts_check) >= 2 and len(parts_check[0].rstrip('.')) <= 2:
            # C'est une initiale (ex: "J. Smith")
            initial = parts_check[0].rstrip('.').upper()
            last_name = ' '.join(parts_check[1:])

            # Chercher dans le mapping si on a déjà un nom complet avec cette initiale et ce nom
            for existing_original, existing_anon in anonymization_map.items():
                existing_parts = existing_original.split()
                if len(existing_parts) >= 2:
                    existing_first = existing_parts[0]
                    existing_last = ' '.join(existing_parts[1:])

                    # Vérifier si l'initiale et le nom correspondent
                    if existing_first[0].upper() == initial and existing_last.lower() == last_name.lower():
                        # Trouvé! Utiliser le même alias mais en format initiale
                        potential_match = existing_anon
                        # Extraire juste l'initiale de l'alias
                        anon_parts = potential_match.split()
                        if len(anon_parts) >= 2:
                            anonymized = f"{anon_parts[0][0]}. {' '.join(anon_parts[1:])}"
                            anonymization_map[original_value] = anonymized
                            return anonymized
                        break

        # Seed basé sur le nom normalisé
        random.seed(hash(normalized_for_hash))

        # Noms masculins professionnels (pour executives et documents SEC)
        male_first_names = [
            'James', 'John', 'Robert', 'Michael', 'William', 'David', 'Richard', 'Joseph',
            'Thomas', 'Charles', 'Christopher', 'Daniel', 'Matthew', 'Anthony', 'Donald',
            'Mark', 'Paul', 'Steven', 'Andrew', 'Kenneth', 'Joshua', 'Kevin', 'Brian',
            'George', 'Edward', 'Ronald', 'Timothy', 'Jason', 'Jeffrey', 'Ryan', 'Jacob',
            'Gary', 'Nicholas', 'Eric', 'Jonathan', 'Stephen', 'Larry', 'Justin', 'Scott',
            'Brandon', 'Benjamin', 'Samuel', 'Raymond', 'Gregory', 'Frank', 'Alexander',
            'Patrick', 'Jack', 'Dennis', 'Jerry', 'Tyler', 'Aaron', 'Henry', 'Douglas',
            'Peter', 'Walter', 'Nathan', 'Zachary', 'Kyle', 'Harold', 'Carl', 'Keith'
        ]

        # Noms féminins professionnels
        female_first_names = [
            'Mary', 'Patricia', 'Jennifer', 'Linda', 'Elizabeth', 'Barbara', 'Susan',
            'Jessica', 'Sarah', 'Karen', 'Nancy', 'Lisa', 'Margaret', 'Betty', 'Sandra',
            'Ashley', 'Dorothy', 'Kimberly', 'Emily', 'Donna', 'Michelle', 'Carol', 'Amanda',
            'Melissa', 'Deborah', 'Stephanie', 'Rebecca', 'Laura', 'Sharon', 'Cynthia',
            'Kathleen', 'Amy', 'Angela', 'Shirley', 'Anna', 'Brenda', 'Pamela', 'Nicole',
            'Ruth', 'Katherine', 'Samantha', 'Christine', 'Catherine', 'Virginia', 'Debra',
            'Rachel', 'Janet', 'Emma', 'Carolyn', 'Maria', 'Heather', 'Diane', 'Julie',
            'Joyce', 'Evelyn', 'Joan', 'Victoria', 'Kelly', 'Christina', 'Lauren', 'Frances'
        ]

        # Noms de famille diversifiés (plus réalistes pour contexte professionnel)
        last_names = [
            'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller', 'Davis',
            'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Wilson', 'Anderson', 'Thomas',
            'Taylor', 'Moore', 'Jackson', 'Martin', 'Lee', 'White', 'Harris', 'Clark',
            'Lewis', 'Robinson', 'Walker', 'Young', 'Allen', 'King', 'Wright', 'Scott',
            'Torres', 'Nguyen', 'Hill', 'Flores', 'Green', 'Adams', 'Nelson', 'Baker',
            'Hall', 'Rivera', 'Campbell', 'Mitchell', 'Carter', 'Roberts', 'Gomez', 'Phillips',
            'Evans', 'Turner', 'Diaz', 'Parker', 'Cruz', 'Edwards', 'Collins', 'Reyes',
            'Stewart', 'Morris', 'Morales', 'Murphy', 'Cook', 'Rogers', 'Gutierrez', 'Ortiz',
            'Morgan', 'Cooper', 'Peterson', 'Bailey', 'Reed', 'Kelly', 'Howard', 'Ramos',
            'Kim', 'Cox', 'Ward', 'Richardson', 'Watson', 'Brooks', 'Chavez', 'Wood',
            'James', 'Bennett', 'Gray', 'Mendoza', 'Ruiz', 'Hughes', 'Price', 'Alvarez',
            'Castillo', 'Sanders', 'Patel', 'Myers', 'Long', 'Ross', 'Foster', 'Jimenez'
        ]

        # Combiner les listes de prénoms
        all_first_names = male_first_names + female_first_names

        # Détecter la structure du nom original
        parts = base_value.split()

        # Détecter suffixes (Jr., Sr., III, etc.)
        suffix = ''
        if len(parts) > 0:
            last_part = parts[-1].rstrip('.,')
            if last_part.upper() in ['JR', 'SR', 'II', 'III', 'IV', 'V', 'ESQ']:
                suffix = f" {last_part}"
                parts = parts[:-1]  # Enlever le suffixe pour le traitement

        if len(parts) >= 2:
            # Check si première partie est une initiale
            if len(parts[0].rstrip('.')) <= 2:
                # Initiale + Nom (ex: "J. Smith")
                first_initial = random.choice(all_first_names)[0]
                last = random.choice(last_names)
                anonymized = f"{first_initial}. {last}{suffix}"

            elif len(parts) == 3 and len(parts[1].rstrip('.')) <= 2:
                # Prénom Initiale Nom (ex: "John A. Smith")
                first = random.choice(all_first_names)
                middle_initial = random.choice(all_first_names)[0]
                last = random.choice(last_names)
                anonymized = f"{first} {middle_initial}. {last}{suffix}"

            elif len(parts) == 3 and len(parts[0].rstrip('.')) <= 2 and len(parts[1].rstrip('.')) <= 2:
                # Initiale Initiale Nom (ex: "J. A. Smith")
                first_initial = random.choice(all_first_names)[0]
                middle_initial = random.choice(all_first_names)[0]
                last = random.choice(last_names)
                anonymized = f"{first_initial}. {middle_initial}. {last}{suffix}"

            else:
                # Prénom Nom standard (ex: "John Smith")
                first = random.choice(all_first_names)
                last = random.choice(last_names)
                anonymized = f"{first} {last}{suffix}"

        elif len(parts) == 1:
            # Juste un nom de famille
            anonymized = random.choice(last_names)

        else:
            # Fallback
            anonymized = f"{random.choice(all_first_names)} {random.choice(last_names)}{suffix}"

        # Préserver la capitalisation (ALL CAPS si original en ALL CAPS)
        if base_value.isupper():
            anonymized = anonymized.upper()

        # REVERSE CHECK: Si on vient de générer "John Smith" et que "J. Smith" existe déjà,
        # utiliser l'alias de "J. Smith" pour cohérence
        if len(parts) >= 2 and len(parts[0]) > 2:  # Nom complet (pas initiale)
            first_name = parts[0]
            last_name_parts = parts[1:]

            # Chercher si "J. Smith" (ou l'initiale) existe déjà
            initial_version = f"{first_name[0]}. {' '.join(last_name_parts)}"

            for existing_original, existing_anon in anonymization_map.items():
                if existing_original.lower() == initial_version.lower():
                    # Trouvé! Utiliser la version complète de cet alias
                    anon_parts = existing_anon.split()
                    if len(anon_parts) >= 2 and len(anon_parts[0].rstrip('.')) <= 2:
                        # L'alias existant est en format initiale, le convertir en nom complet
                        # On doit retrouver le nom complet à partir de l'initiale
                        # Pour cela, on utilise le même seed que l'original
                        temp_seed = hash(initial_version.lower())
                        random.seed(temp_seed)
                        all_names_temp = male_first_names + female_first_names
                        # Trouver un prénom qui commence par la bonne initiale
                        matching_first = [n for n in all_names_temp if n[0].upper() == anon_parts[0][0].upper()]
                        if matching_first:
                            # Utiliser le premier match avec le bon seed
                            random.seed(temp_seed)
                            full_first = random.choice(matching_first)
                            anonymized = f"{full_first} {' '.join(anon_parts[1:])}{suffix}"
                            break

    elif info_type == 'email':
        # Générer email anonymisé
        parts = original_value.split('@')
        if len(parts) == 2:
            local_part = parts[0]
            domain = parts[1]

            # Anonymiser le domaine si c'est un domaine d'entreprise
            if domain in anonymization_map:
                anon_domain = anonymization_map[domain].replace('http://', '').replace('https://', '').replace('www.', '')
            else:
                random.seed(hash(domain))
                # Noms de domaines réalistes
                company_domains = [
                    'riverside-capital.com', 'oakmont-financial.com', 'clearwater-holdings.com',
                    'bridgeport-advisors.com', 'summit-partners.com', 'westfield-group.com',
                    'highland-investments.com', 'parkview-associates.com', 'greenfield-equity.com',
                    'sterling-capital.com', 'premier-financial.com', 'meridian-partners.com',
                    'cornerstone-holdings.com', 'heritage-advisors.com', 'beacon-group.com'
                ]
                anon_domain = random.choice(company_domains)
                anonymization_map[domain] = anon_domain

            # Anonymiser la partie locale avec des noms réalistes
            random.seed(hash(local_part))
            # Format: prénom.nom
            first_names = ['john', 'jane', 'michael', 'sarah', 'david', 'emily', 'robert', 'lisa',
                          'james', 'jennifer', 'william', 'patricia', 'thomas', 'linda', 'charles',
                          'barbara', 'daniel', 'susan', 'matthew', 'jessica', 'andrew', 'karen']
            last_names = ['smith', 'johnson', 'williams', 'brown', 'jones', 'miller', 'davis',
                         'garcia', 'rodriguez', 'wilson', 'martinez', 'anderson', 'taylor',
                         'thomas', 'moore', 'jackson', 'martin', 'lee', 'thompson', 'white']

            # Déterminer le format de l'email original
            if '.' in local_part:
                # Format: prénom.nom
                anon_local = f"{random.choice(first_names)}.{random.choice(last_names)}"
            elif '_' in local_part:
                # Format: prénom_nom
                anon_local = f"{random.choice(first_names)}_{random.choice(last_names)}"
            else:
                # Format simple: juste un nom
                anon_local = random.choice(last_names)

            anonymized = f"{anon_local}@{anon_domain}"

    elif info_type == 'phone':
        # Randomiser les chiffres en gardant le format
        random.seed(hash(original_value))
        anonymized = re.sub(r'\d', lambda m: str(random.randint(0, 9)), original_value)

    elif info_type == 'website':
        # Anonymiser les sites web en conservant la structure
        random.seed(hash(original_value))

        # Extraire le protocole (http://, https://) et le préfixe www.
        protocol = ''
        if original_value.startswith('https://'):
            protocol = 'https://'
            domain = original_value[8:]
        elif original_value.startswith('http://'):
            protocol = 'http://'
            domain = original_value[7:]
        else:
            domain = original_value

        # Extraire www. si présent
        www_prefix = ''
        if domain.startswith('www.'):
            www_prefix = 'www.'
            domain = domain[4:]

        # Séparer le domaine principal et le path
        path = ''
        if '/' in domain:
            parts = domain.split('/', 1)
            domain = parts[0]
            path = '/' + parts[1]

        # Générer un domaine anonymisé réaliste
        domain_prefixes = [
            'riverside', 'oakmont', 'clearwater', 'bridgeport', 'summit', 'westfield',
            'highland', 'parkview', 'greenfield', 'sterling', 'premier', 'meridian',
            'cornerstone', 'heritage', 'beacon', 'cascade', 'horizon', 'venture'
        ]
        domain_types = [
            'capital', 'financial', 'holdings', 'group', 'partners', 'advisors',
            'services', 'solutions', 'systems', 'tech', 'digital', 'consulting'
        ]

        # Extraire le TLD (top-level domain) de l'original
        tld = 'com'  # Default
        if '.' in domain:
            tld = domain.split('.')[-1]

        # Générer le nouveau domaine
        prefix = random.choice(domain_prefixes)
        dtype = random.choice(domain_types)
        anon_domain = f"{prefix}-{dtype}.{tld}"

        # Reconstituer l'URL anonymisée
        anonymized = f"{protocol}{www_prefix}{anon_domain}{path}"

    elif info_type in ['address', 'partial_address']:
        # Générer adresse aléatoire ou partielle
        random.seed(hash(base_value))
        street_nums = [str(random.randint(100, 999))]
        street_names = ['Main Street', 'Oak Avenue', 'Maple Drive', 'Cedar Lane', 'Pine Road',
                      'Elm Street', 'Park Avenue', 'Washington Boulevard', 'Lincoln Way', 'Market Street']
        cities = ['New York', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix', 'Philadelphia',
                 'San Antonio', 'San Diego', 'Dallas', 'San Jose', 'Boston', 'Seattle', 'Denver']
        states = ['NY', 'CA', 'TX', 'FL', 'IL', 'PA', 'OH', 'GA', 'NC', 'MI', 'NJ', 'VA', 'WA', 'MA']
        zip_code = ''.join([str(random.randint(0, 9)) for _ in range(5)])

        # Détecter le format de l'adresse partielle
        if ',' in base_value:
            # Format avec virgule (ex: "Edgewood, NY 11717" ou "Washington, DC 20549")
            parts = base_value.split(',')
            if len(parts) >= 2:
                # Générer format similaire
                city = random.choice(cities)
                state = random.choice(states)
                anonymized = f"{city}, {state} {zip_code}"
            else:
                anonymized = f"{random.choice(street_nums)} {random.choice(street_names)}, {random.choice(cities)} {zip_code}"
        else:
            # Adresse complète ou autre format
            anonymized = f"{random.choice(street_nums)} {random.choice(street_names)}, {random.choice(cities)} {zip_code}"

    elif info_type == 'city_name':
        # Générer nom de ville aléatoire
        random.seed(hash(base_value))
        cities = ['Springfield', 'Riverside', 'Centerville', 'Georgetown', 'Franklin', 'Clinton',
                 'Madison', 'Washington', 'Arlington', 'Manchester', 'Oxford', 'Cambridge', 'Salem']
        anonymized = random.choice(cities)

    elif info_type == 'ssn':
        # Randomiser SSN en gardant le format XXX-XX-XXXX
        random.seed(hash(original_value))
        anonymized = f"{random.randint(100, 999):03d}-{random.randint(10, 99):02d}-{random.randint(1000, 9999):04d}"

    elif info_type == 'credit_card':
        # Randomiser numéro de carte en gardant le format
        random.seed(hash(original_value))
        anonymized = re.sub(r'\d', lambda m: str(random.randint(0, 9)), original_value)

    elif info_type in ['cusip_code', 'isin_code']:
        # Randomiser codes en gardant le format
        random.seed(hash(original_value))
        anonymized = re.sub(r'[A-Z0-9]', lambda m: random.choice('ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789') if m.group().isalnum() else m.group(), original_value)

    elif info_type == 'quoted_title':
        # Garder les guillemets, anonymiser le contenu
        random.seed(hash(base_value))
        titles = ['Strategic Initiative', 'Annual Report', 'Market Analysis', 'Financial Overview',
                 'Corporate Strategy', 'Business Plan', 'Performance Review', 'Investment Proposal']
        if '"' in original_value or '"' in original_value:
            anonymized = f'"{random.choice(titles)}"'
        else:
            anonymized = random.choice(titles)

    elif info_type in ['cik_number', 'commission_file_number', 'irs_ein']:
        # Randomiser les numéros en gardant le format
        random.seed(hash(original_value))
        anonymized = re.sub(r'\d', lambda m: str(random.randint(0, 9)), original_value)

    elif info_type == 'ticker_symbol':
        # Générer ticker aléatoire
        random.seed(hash(base_value))
        letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        ticker_len = len(re.sub(r'[^A-Z]', '', base_value.upper()))
        anonymized = ''.join(random.choice(letters) for _ in range(ticker_len))
        # Préserver le format (ex: (NYSE: XXXX))
        if '(' in original_value:
            match = re.match(r'\(([^:]+):\s*', original_value)
            if match:
                exchange = match.group(1)
                anonymized = f"({exchange}: {anonymized})"

    elif info_type == 'important_date':
        # Randomiser les dates en gardant le format
        random.seed(hash(original_value))
        anonymized = re.sub(r'\d', lambda m: str(random.randint(0, 9)), original_value)

    elif info_type in ['auditor', 'underwriter', 'law_firm', 'financial_institution', 'professional_firm']:
        # Générateur de noms de firmes professionnelles réalistes
        random.seed(hash(base_value))

        # Noms pour cabinets d'audit (style Big 4)
        if info_type == 'auditor':
            firm_patterns = [
                # Pattern 1: Nom + "& Young/& Co/& Associates"
                ('Anderson', 'Carter', 'Edwards', 'Harrison', 'Morrison', 'Patterson', 'Wellington'),
                (' & Young LLP', ' & Co. LLP', ' & Associates LLP', ' & Partners LLP'),
                # Pattern 2: Double nom (style "Ernst & Young")
                ('Sterling', 'Reynolds', 'Davidson', 'Franklin', 'Lawrence', 'Thompson'),
                (' & Young LLP', ' & Partners LLP', ' & Touche LLP')
            ]
            name1 = random.choice(firm_patterns[0])
            suffix = random.choice(firm_patterns[1])
            anonymized = f"{name1}{suffix}"

        # Noms pour cabinets d'avocats
        elif info_type == 'law_firm':
            name_parts = ['Anderson', 'Bennett', 'Carter', 'Davidson', 'Edwards', 'Franklin',
                         'Harrison', 'Kennedy', 'Morrison', 'Patterson', 'Reynolds', 'Sterling',
                         'Thompson', 'Wallace', 'Wellington', 'Whitman']
            # 3-4 noms séparés par des virgules (style cabinet d'avocats)
            num_names = random.randint(2, 4)
            selected_names = random.sample(name_parts, num_names)
            if num_names == 2:
                anonymized = f"{selected_names[0]} & {selected_names[1]} LLP"
            else:
                anonymized = ', '.join(selected_names[:-1]) + f" & {selected_names[-1]} LLP"

        # Noms pour underwriters (banques d'investissement)
        elif info_type == 'underwriter':
            bank_names = [
                'Sterling Securities LLC', 'Riverside Capital Markets LLC', 'Highland Securities LLC',
                'Meridian Investment Banking LLC', 'Summit Capital Markets LLC',
                'Cornerstone Securities LLC', 'Heritage Investment Banking LLC',
                'Oakmont Securities LLC', 'Clearwater Capital LLC', 'Bridgeport Securities LLC'
            ]
            anonymized = random.choice(bank_names)

        # Noms pour institutions financières (banques, trustees)
        elif info_type == 'financial_institution':
            institution_types = ['Bank', 'Trust Company', 'National Bank', 'Bancorp', 'Credit Union']
            city_names = ['Riverside', 'Oakmont', 'Clearwater', 'Summit', 'Highland', 'Westfield',
                         'Parkview', 'Greenfield', 'Maplewood', 'Brookfield', 'Hillcrest', 'Bayshore']
            city = random.choice(city_names)
            inst_type = random.choice(institution_types)
            if 'National' in inst_type or 'Bancorp' in inst_type:
                anonymized = f"{city} {inst_type}"
            else:
                anonymized = f"{city} {inst_type}, N.A."

        # Autres firmes professionnelles
        else:
            firm_styles = [
                'Anderson & Associates', 'Carter Partners LLP', 'Edwards Group',
                'Harrison & Co. LLP', 'Morrison Partners', 'Reynolds & Associates',
                'Sterling Group LLP', 'Thompson & Partners', 'Wellington Associates'
            ]
            anonymized = random.choice(firm_styles)

    else:
        # Par défaut, masquer avec des X
        anonymized = 'X' * min(len(base_value), 10)

    # Sauvegarder le mapping
    anonymization_map[original_value] = anonymized

    return anonymized


# ===== DEDUPLICATION FUNCTIONS =====

def _clean_name_for_deduplication(name: str, is_person: bool = False, is_firm: bool = False) -> str:
    """
    Clean a name for deduplication by removing garbage characters, numbers, and normalizing.

    Args:
        name: Original name string
        is_person: True if person name (applies extra cleaning)
        is_firm: True if firm/company name (normalizes punctuation)

    Returns:
        Cleaned, normalized name for comparison
    """
    cleaned = name.lower().strip()
    cleaned = re.sub(r'\s+', ' ', cleaned)

    if is_firm:
        # Normalize punctuation for firm/company names
        # "Ernst & Young, LLP" → "ernst & young llp"
        # "Ernst & Young LLP" → "ernst & young llp"
        cleaned = re.sub(r',\s*', ' ', cleaned)  # Remove commas
        cleaned = re.sub(r'\.', '', cleaned)      # Remove periods from "L.L.P." → "llp"
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()

    if is_person:
        # Remove dates/numbers at beginning: "2/17/2026 S. Ferris" → "s. ferris"
        cleaned = re.sub(r'^[\d/,\s]+(?=[a-z])', '', cleaned)

        # Remove numbers/dates at end: "Jeffrey Stieﬂer 138,750" → "jeffrey stieﬂer"
        cleaned = re.sub(r'\s+[\d/,]+$', '', cleaned)

        # Remove parentheses with numbers at end: "Charles Drucker(7)(8" → "charles drucker"
        # This pattern handles multiple groups: (7)(8, (7)(8), (15, etc.
        # Keep applying until no more matches
        while re.search(r'\s*\([^\)]*\d[^\)]*\)*$', cleaned):
            cleaned = re.sub(r'\s*\([^\)]*\d[^\)]*\)*$', '', cleaned)
        while re.search(r'\s*\(\d+[^\)]*$', cleaned):
            cleaned = re.sub(r'\s*\(\d+[^\)]*$', '', cleaned)

        # Remove trailing dashes: "David Karnstedt - - -" → "david karnstedt"
        cleaned = re.sub(r'[\s\-]+$', '', cleaned)

        # Remove signature markers: "/s/", "(s)", "signed:", etc.
        cleaned = re.sub(r'\s*/[sS]/\s*', ' ', cleaned)
        cleaned = re.sub(r'\s*\([sS]\)\s*', ' ', cleaned)
        cleaned = re.sub(r'\s+signed:?\s*', ' ', cleaned, flags=re.IGNORECASE)

        # Remove middle initials: "stephanie l. ferris" → "stephanie ferris"
        cleaned = re.sub(r'\s+[a-z]\.\s+', ' ', cleaned)  # "john a. smith" → "john smith"
        cleaned = re.sub(r'\s+[a-z]\s+', ' ', cleaned)    # "john a smith" → "john smith"

    # Final cleanup
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned


def _deduplicate_with_prefix_matching(items_dict, info_type):
    """
    Deduplicates items using bidirectional prefix matching.
    For company names/addresses: "Vantiv" matches "Vantiv LLC"
    For person names: "Charles D. Drucker" matches "Charles Drucker" (ignores middle initials)
    For firms: "Ernst & Young LLP" matches "Ernst & Young, LLP" (handles punctuation)
    Keeps shortest/cleanest version.

    Args:
        items_dict: Dictionary of (type, value) -> [pages]
        info_type: Type to deduplicate

    Returns:
        Deduplicated dictionary
    """
    # Types that benefit from deduplication
    dedup_types = [
        'company_name', 'address', 'partial_address',
        'person_name', 'executive_name',
        'auditor', 'underwriter', 'law_firm', 'financial_institution', 'professional_firm'
    ]

    if info_type not in dedup_types:
        return items_dict  # Skip deduplication for other types

    deduplicated = {}

    for (item_type, item_value), pages in items_dict.items():
        if item_type != info_type:
            deduplicated[(item_type, item_value)] = pages
            continue

        # Clean and normalize the value
        is_person = (info_type in ['person_name', 'executive_name'])
        is_firm = (info_type in ['company_name', 'auditor', 'underwriter', 'law_firm',
                                   'financial_institution', 'professional_firm'])
        normalized = _clean_name_for_deduplication(item_value, is_person=is_person, is_firm=is_firm)

        matching_key = None

        # Check if this value is a prefix/suffix of existing entries
        for existing_key in list(deduplicated.keys()):
            if existing_key[0] != info_type:
                continue

            existing_value = existing_key[1]
            # Clean the existing value the same way
            existing_normalized = _clean_name_for_deduplication(existing_value, is_person=is_person, is_firm=is_firm)

            # EXACT match after normalization
            if normalized == existing_normalized:
                matching_key = existing_key
                break

            # Bidirectional PREFIX check
            # Ex: "Lisa Hook" matches "Lisa Hook BOON"
            # Ex: "Vantiv" matches "Vantiv LLC"
            if existing_normalized.startswith(normalized + ' ') or \
               normalized.startswith(existing_normalized + ' '):
                matching_key = existing_key
                break

        if matching_key:
            # Merge pages
            deduplicated[matching_key].extend(pages)

            # Keep the SHORTEST value (usually the cleanest)
            # Ex: "Charles Drucker" is kept over "Charles Drucker(7)(8"
            # Ex: "Lisa Hook" is kept over "Lisa Hook /s/ BOON"
            if len(item_value) < len(matching_key[1]):
                # Replace with shorter version
                deduplicated[(item_type, item_value)] = deduplicated.pop(matching_key)

            # SPECIAL: For person names, prefer proper case over ALL CAPS
            # Ex: "John Smith" is kept over "JOHN SMITH"
            elif info_type in ['person_name', 'executive_name']:
                current_value = matching_key[1]
                # If existing is ALL CAPS but new one is NOT, replace with new
                if current_value.isupper() and not item_value.isupper():
                    deduplicated[(item_type, item_value)] = deduplicated.pop(matching_key)
        else:
            deduplicated[(item_type, item_value)] = pages

    return deduplicated


# ===== BATCH CONSOLIDATION FUNCTIONS =====

def _consolidate_batch_findings(all_findings_by_pdf, all_images_by_pdf):
    """
    Consolidate findings from multiple PDFs, deduplicating by (type, value).

    Args:
        all_findings_by_pdf: {
            'file1.pdf': {page_num: [{type, value, page}, ...]},
            'file2.pdf': {page_num: [{type, value, page}, ...]},
            ...
        }
        all_images_by_pdf: {
            'file1.pdf': {page_num: count},
            'file2.pdf': {page_num: count},
            ...
        }

    Returns:
        consolidated: {
            ('email', 'user@example.com'): {
                'type': 'email',
                'value': 'user@example.com',
                'occurrences': [
                    {'document': 'file1.docx', 'pages': [5, 12]},
                    {'document': 'file2.docx', 'pages': [3]}
                ]
            },
            ...
        }
    """
    consolidated = {}

    # Process sensitive information findings
    for pdf_filename, findings_by_page in all_findings_by_pdf.items():
        # Convert pdf filename to docx for display
        doc_name = pdf_filename.replace('.pdf', '.docx')

        for page_num, findings_list in findings_by_page.items():
            for finding in findings_list:
                info_type = finding.get('type', '')
                info_value = finding.get('value', '')
                page = finding.get('page', page_num)

                # Normalize value for better deduplication
                normalized_value = info_value
                if info_type == 'email':
                    normalized_value = info_value.lower()  # Emails: lowercase
                elif info_type in ['person_name', 'executive_name', 'company_name']:
                    # Names: prefer proper case over ALL CAPS
                    if info_value.isupper() and len(info_value) > 3:
                        # If ALL CAPS, check if we have a better version
                        pass  # Will be handled in merge logic
                    else:
                        normalized_value = info_value

                # Create unique key (type, normalized_value)
                key = (info_type, normalized_value)

                if key not in consolidated:
                    consolidated[key] = {
                        'type': info_type,
                        'value': info_value,  # Keep original formatting
                        'occurrences': []
                    }

                # Find if this document already has an entry
                existing_doc = None
                for occ in consolidated[key]['occurrences']:
                    if occ['document'] == doc_name:
                        existing_doc = occ
                        break

                if existing_doc:
                    # Add page to existing document entry
                    if page not in existing_doc['pages']:
                        existing_doc['pages'].append(page)
                        existing_doc['pages'].sort()
                else:
                    # Create new document entry
                    consolidated[key]['occurrences'].append({
                        'document': doc_name,
                        'pages': [page]
                    })

                # Update value to prefer better formatting
                if info_type in ['person_name', 'executive_name', 'company_name']:
                    current_value = consolidated[key]['value']
                    # Prefer proper case over ALL CAPS
                    if current_value.isupper() and not info_value.isupper():
                        consolidated[key]['value'] = info_value
                    # Prefer longer version if same case
                    elif len(info_value) > len(current_value):
                        if info_value.lower() == current_value.lower():
                            consolidated[key]['value'] = info_value

    # Process images
    for pdf_filename, images_by_page in all_images_by_pdf.items():
        doc_name = pdf_filename.replace('.pdf', '.docx')

        for page_num, image_count in images_by_page.items():
            if image_count > 0:
                info_type = 'image'
                info_value = f"{image_count} image(s)"

                key = (info_type, info_value)

                if key not in consolidated:
                    consolidated[key] = {
                        'type': info_type,
                        'value': info_value,
                        'occurrences': []
                    }

                # Find if this document already has an entry
                existing_doc = None
                for occ in consolidated[key]['occurrences']:
                    if occ['document'] == doc_name:
                        existing_doc = occ
                        break

                if existing_doc:
                    if page_num not in existing_doc['pages']:
                        existing_doc['pages'].append(page_num)
                        existing_doc['pages'].sort()
                else:
                    consolidated[key]['occurrences'].append({
                        'document': doc_name,
                        'pages': [page_num]
                    })

    return consolidated


def _create_consolidated_excel_report(consolidated_findings, report_path, grouping_keywords=None):
    """
    Create consolidated Excel report from findings across multiple PDFs.

    Args:
        consolidated_findings: Dict from _consolidate_batch_findings()
        report_path: Path to save the Excel file
        grouping_keywords: Optional list of keywords for grouping
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available, cannot create Excel report")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Analysis"

    # Headers
    headers = ["Information Type", "Information", "", "Documents", "Pages", "Anonymized Information"]
    ws.append(headers)

    # Dictionnaire pour maintenir la cohérence des anonymisations
    anonymization_map = {}

    # Style headers
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Type labels (same as in create_excel_report)
    type_labels = {
        'email': 'Email',
        'phone': 'Phone',
        'address': 'Address',
        'partial_address': 'Partial Address',
        'city_name': 'City Name',
        'ssn': 'SSN',
        'credit_card': 'Credit Card',
        'irs_ein': 'IRS EIN',
        'commission_file_number': 'Commission File Number',
        'cik_number': 'CIK Number',
        'cusip_code': 'CUSIP Code',
        'isin_code': 'ISIN Code',
        'executive_name': 'Executive Name',
        'person_name': 'Person Name',
        'company_name': 'Company Name',
        'ticker_symbol': 'Ticker Symbol',
        'important_date': 'Important Date',
        'auditor': 'Auditor',
        'underwriter': 'Underwriter',
        'law_firm': 'Law Firm',
        'financial_institution': 'Financial Institution',
        'professional_firm': 'Professional Firm',
        'quoted_title': 'Quoted Title',
        'image': 'Image'
    }

    # Color mapping (same as in create_excel_report)
    color_map = {
        'email': 'E3F2FD',
        'phone': 'E8F5E9',
        'address': 'FFF9C4',
        'partial_address': 'FFF9C4',
        'person_name': 'F3E5F5',
        'executive_name': 'F3E5F5',
        'company_name': 'FFE0B2',
        'ssn': 'FFCDD2',
        'credit_card': 'FFCDD2',
        'irs_ein': 'E1F5FE',
        'commission_file_number': 'E1F5FE',
        'cik_number': 'E1F5FE',
        'cusip_code': 'E1F5FE',
        'isin_code': 'E1F5FE',
        'ticker_symbol': 'E1F5FE',
        'image': 'F5F5F5'
    }

    # Prepare data for Excel
    rows_data = []

    # GROUPING BY KEYWORDS (if provided)
    # Group findings that contain the same keyword
    if grouping_keywords:
        grouped_findings = {}  # {(type, keyword): {value, occurrences}}

        for (info_type, _), finding_data in consolidated_findings.items():
            info_value = finding_data['value']
            info_value_lower = info_value.lower()

            # Check if this value contains any of the keywords
            matched_keyword = None
            for keyword in grouping_keywords:
                if keyword.lower() in info_value_lower:
                    matched_keyword = keyword
                    break

            if matched_keyword:
                # Group under the keyword
                key = (info_type, matched_keyword.lower())
                if key not in grouped_findings:
                    grouped_findings[key] = {
                        'type': info_type,
                        'value': matched_keyword,  # Display keyword as the value
                        'occurrences': []
                    }
                # Merge occurrences
                for occ in finding_data['occurrences']:
                    # Check if document already exists
                    existing = None
                    for existing_occ in grouped_findings[key]['occurrences']:
                        if existing_occ['document'] == occ['document']:
                            existing = existing_occ
                            break

                    if existing:
                        # Merge pages
                        for page in occ['pages']:
                            if page not in existing['pages']:
                                existing['pages'].append(page)
                        existing['pages'].sort()
                    else:
                        # Add new document entry
                        grouped_findings[key]['occurrences'].append({
                            'document': occ['document'],
                            'pages': occ['pages'][:]
                        })
            else:
                # Not grouped - keep original
                key = (info_type, info_value.lower())
                grouped_findings[key] = finding_data

        # Use grouped findings
        consolidated_findings = grouped_findings

    for (info_type, _), finding_data in consolidated_findings.items():
        info_value = finding_data['value']
        occurrences = finding_data['occurrences']

        # Column A: Type label
        type_label = type_labels.get(info_type, info_type.replace('_', ' ').title())

        # Column B: Information value
        value_str = str(info_value)

        # Column D: Documents (comma-separated)
        documents = [occ['document'] for occ in occurrences]
        documents_str = ', '.join(documents)

        # Column E: Pages (grouped by document)
        pages_parts = []
        for occ in occurrences:
            doc_name = occ['document']
            pages = occ['pages']
            pages_str = ', '.join(f"p.{p}" for p in sorted(pages))
            pages_parts.append(f"{doc_name} {pages_str}")

        pages_str = '; '.join(pages_parts)

        rows_data.append({
            'type': info_type,
            'type_label': type_label,
            'value': value_str,
            'documents': documents_str,
            'pages': pages_str,
            'anonymized': ''  # Will be filled when creating rows
        })

    # Sort by type, then by value
    rows_data.sort(key=lambda x: (x['type_label'], x['value']))

    # Write rows to Excel
    for row_data in rows_data:
        # Generate anonymized value
        anonymized_value = _anonymize_value(row_data['value'], row_data['type'], anonymization_map)

        row = [
            row_data['type_label'],  # Column A
            row_data['value'],        # Column B
            '',                        # Column C (empty spacer)
            row_data['documents'],     # Column D
            row_data['pages'],         # Column E
            anonymized_value           # Column F
        ]
        ws.append(row)

        # Apply color coding
        row_num = ws.max_row
        info_type = row_data['type']
        if info_type in color_map:
            fill_color = color_map[info_type]
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            for col in range(1, 7):  # Columns A-F
                ws.cell(row=row_num, column=col).fill = fill

    # Set column widths
    ws.column_dimensions['A'].width = 25  # Information Type
    ws.column_dimensions['B'].width = 50  # Information
    ws.column_dimensions['C'].width = 5   # Empty spacer
    ws.column_dimensions['D'].width = 40  # Documents
    ws.column_dimensions['E'].width = 60  # Pages (longer for multi-document)
    ws.column_dimensions['F'].width = 50  # Anonymized Information

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save
    wb.save(str(report_path))
    logger.info(f"Consolidated report saved to {report_path}")

    return report_path


def create_consolidated_batch_report(all_findings_by_pdf, all_images_by_pdf, output_path, grouping_keywords=None):
    """
    Public wrapper to create a consolidated batch report.

    Args:
        all_findings_by_pdf: Dict mapping PDF filenames to findings by page
        all_images_by_pdf: Dict mapping PDF filenames to images by page
        output_path: Path where to save the consolidated Excel report
        grouping_keywords: Optional list of keywords for grouping similar entries

    Returns:
        Path to the created report, or None if failed
    """
    # Consolidate findings from all PDFs
    consolidated = _consolidate_batch_findings(all_findings_by_pdf, all_images_by_pdf)

    # Create the Excel report
    return _create_consolidated_excel_report(consolidated, output_path, grouping_keywords)


# ===== REPORT GENERATION FUNCTIONS =====

def analyze_and_create_report(
    pdf_path: Path,
    report_path: Optional[Path] = None,
    grouping_keywords: Optional[List[str]] = None,
    verbose: bool = False
) -> Path:
    """
    Analyse le PDF et crée un rapport.
    
    Args:
        pdf_path: Chemin vers le PDF
        report_path: Chemin du fichier de rapport (optionnel)
        grouping_keywords: Liste de mots-clés pour regrouper les occurrences
        verbose: Si True, affiche les logs détaillés
    
    Returns:
        Path: Chemin du rapport créé
    """
    if verbose:
        logger.info("Starting PDF analysis...")
    
    # Détecter les images
    images_by_page = detect_images_in_pdf(pdf_path)
    if verbose:
        logger.info(f"Images detected on {len(images_by_page)} page(s)")
    
    # Détecter les informations sensibles
    sensitive_info_by_page = detect_sensitive_information(pdf_path, verbose=verbose)
    
    total_sensitive_pages = len(sensitive_info_by_page)
    logger.info(f"Sensitive information detected on {total_sensitive_pages} page(s)")
    
    # Créer le chemin du rapport
    if report_path is None:
        report_path = pdf_path.parent / f"{pdf_path.stem}_ANALYSE.xlsx"
    else:
        report_path = Path(report_path)
    
    # Créer le rapport
    if OPENPYXL_AVAILABLE:
        create_excel_report(
            report_path, pdf_path, images_by_page, sensitive_info_by_page, grouping_keywords
        )
    else:
        create_text_report(
            report_path, pdf_path, images_by_page, sensitive_info_by_page
        )
    
    logger.info(f"Report created: {report_path}")
    return report_path


def create_excel_report(
    report_path: Path,
    pdf_path: Path,
    images_by_page: Dict[int, int],
    sensitive_info_by_page: Dict[int, List[Dict[str, Any]]],
    grouping_keywords: Optional[List[str]] = None
) -> None:
    """
    Crée un rapport Excel avec la structure demandée.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is not available")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis Report"

    # Headers (added Column F for Anonymized Information)
    headers = ["Information Type", "Information", "", "Document", "Pages", "Anonymized Information"]
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Initialize anonymization map for consistent aliases
    anonymization_map = {}

    # Collecter toutes les informations avec leurs pages
    info_dict = {}  # (type, value) -> [pages]
    
    # Images
    for page_num, image_count in images_by_page.items():
        key = ('image', f"{image_count} image(s)")
        if key not in info_dict:
            info_dict[key] = []
        info_dict[key].append(page_num)
    
    # Informations sensibles
    type_labels = {
        'email': 'Email',
        'phone': 'Phone',
        'address': 'Address',
        'partial_address': 'Partial Address',
        'city_name': 'City Name',
        'ssn': 'SSN',
        'credit_card': 'Credit Card',
        'irs_ein': 'IRS EIN',
        'commission_file_number': 'Commission File Number',
        'cik_number': 'CIK Number',
        'cusip_code': 'CUSIP Code',
        'isin_code': 'ISIN Code',
        'executive_name': 'Executive Name',
        'company_name': 'Company Name',
        'ticker_symbol': 'Ticker Symbol',
        'important_date': 'Important Date',
        'auditor': 'Auditor',
        'underwriter': 'Underwriter',
        'law_firm': 'Law Firm',
        'financial_institution': 'Financial Institution',
        'professional_firm': 'Professional Firm',
        'quoted_title': 'Quoted Title',
        'image': 'Image'
    }
    
    for page_num, findings in sensitive_info_by_page.items():
        for finding in findings:
            info_type = finding['type']
            info_value = finding['value']
            key = (info_type, info_value)
            if key not in info_dict:
                info_dict[key] = []
            info_dict[key].append(page_num)

    # Apply prefix-based deduplication for company names, addresses, person names, and firms
    info_dict = _deduplicate_with_prefix_matching(info_dict, 'company_name')
    info_dict = _deduplicate_with_prefix_matching(info_dict, 'address')
    info_dict = _deduplicate_with_prefix_matching(info_dict, 'partial_address')
    info_dict = _deduplicate_with_prefix_matching(info_dict, 'person_name')
    info_dict = _deduplicate_with_prefix_matching(info_dict, 'executive_name')
    # Professional firm types (handle punctuation variations like "Ernst & Young LLP" vs "Ernst & Young, LLP")
    info_dict = _deduplicate_with_prefix_matching(info_dict, 'auditor')
    info_dict = _deduplicate_with_prefix_matching(info_dict, 'underwriter')
    info_dict = _deduplicate_with_prefix_matching(info_dict, 'law_firm')
    info_dict = _deduplicate_with_prefix_matching(info_dict, 'financial_institution')
    info_dict = _deduplicate_with_prefix_matching(info_dict, 'professional_firm')

    # Regrouper par mots-clés si spécifiés
    if grouping_keywords:
        grouped_dict = {}
        keywords_lower = [kw.lower() for kw in grouping_keywords]
        
        for (info_type, info_value), pages in info_dict.items():
            info_value_lower = str(info_value).lower()
            
            # Chercher si cette valeur contient un mot-clé
            matched_keyword = None
            for keyword in keywords_lower:
                if keyword in info_value_lower:
                    matched_keyword = keyword
                    break
            
            if matched_keyword:
                # Regrouper sous le mot-clé
                grouped_key = (info_type, matched_keyword.title())
                if grouped_key not in grouped_dict:
                    grouped_dict[grouped_key] = []
                grouped_dict[grouped_key].extend(pages)
            else:
                # Garder tel quel
                if (info_type, info_value) not in grouped_dict:
                    grouped_dict[(info_type, info_value)] = []
                grouped_dict[(info_type, info_value)].extend(pages)
        
        info_dict = grouped_dict
    
    # Trier et ajouter les lignes (important_date à la fin)
    normal_items = []
    important_date_items = []
    
    for item in info_dict.items():
        info_type = item[0][0]
        if info_type == 'important_date':
            important_date_items.append(item)
        else:
            normal_items.append(item)
    
    # Trier les éléments normaux
    sorted_normal = sorted(normal_items, key=lambda x: (x[0][0], x[0][1]))
    # Trier les important_date
    sorted_important_dates = sorted(important_date_items, key=lambda x: (x[0][0], x[0][1]))
    
    # Combiner: normaux d'abord, puis important_date à la fin
    sorted_items = sorted_normal + sorted_important_dates
    
    for (info_type, info_value), pages in sorted_items:
        type_label = type_labels.get(info_type, info_type.capitalize())
        pages_sorted = sorted(set(pages))
        pages_str = ', '.join(f"p.{p}" for p in pages_sorted)
        document_name = pdf_path.stem

        # Generate anonymized value
        anonymized_value = _anonymize_value(str(info_value), info_type, anonymization_map)

        # Append row with 6 columns: Type, Information, Empty, Document, Pages, Anonymized
        ws.append([type_label, info_value, "", document_name, pages_str, anonymized_value])

    # Ajuster la largeur des colonnes (added Column F)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 50  # Anonymized Information
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Sauvegarder
    wb.save(report_path)


def create_text_report(
    report_path: Path,
    pdf_path: Path,
    images_by_page: Dict[int, int],
    sensitive_info_by_page: Dict[int, List[Dict[str, Any]]]
) -> None:
    """Crée un rapport texte si openpyxl n'est pas disponible."""
    type_labels = {
        'email': 'Email',
        'phone': 'Phone',
        'address': 'Address',
        'ssn': 'SSN',
        'credit_card': 'Credit Card',
        'irs_ein': 'IRS EIN',
        'commission_file_number': 'Commission File Number',
        'zip_code': 'ZIP Code',
        'executive_name': 'Executive Name',
        'image': 'Image'
    }
    
    info_regrouped = {}
    
    # Images
    for page_num, image_count in images_by_page.items():
        key = ('image', f"{image_count} image(s)")
        if key not in info_regrouped:
            info_regrouped[key] = []
        info_regrouped[key].append(page_num)
    
    # Informations sensibles
    for page_num, findings in sensitive_info_by_page.items():
        for finding in findings:
            info_type = finding['type']
            info_value = finding['value']
            key = (info_type, info_value)
            if key not in info_regrouped:
                info_regrouped[key] = []
            info_regrouped[key].append(page_num)
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("RAPPORT D'ANALYSE DU PDF\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"Fichier analysé: {pdf_path.name}\n")
        f.write(f"Date d'analyse: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        if info_regrouped:
            f.write(f"{'Information Type':<30} {'Information':<40} {'Document and Pages':<30}\n")
            f.write("-" * 80 + "\n")
            
            sorted_items = sorted(info_regrouped.items(), key=lambda x: (x[0][0], x[0][1]))
            
            for (info_type, info_value), pages in sorted_items:
                type_str = type_labels.get(info_type, info_type)
                value_str = str(info_value)[:38]
                pages_sorted = sorted(set(pages))
                pages_str = ', '.join(f"p.{p}" for p in pages_sorted)
                document_name = pdf_path.stem
                pages_info = f"{document_name} {pages_str}"
                
                f.write(f"{type_str:<30} {value_str:<40} {pages_info:<30}\n")
        else:
            f.write("No images or sensitive information detected.\n")
